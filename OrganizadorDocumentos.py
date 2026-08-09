import os, re, json, shutil, unicodedata, difflib, threading, hashlib, hmac
import smtplib
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from rapidfuzz import fuzz, process
from datetime import datetime, timezone
from tqdm import tqdm
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.utils import encode_rfc2231

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    DND_AVAILABLE = True
except ImportError:
    DND_AVAILABLE = False
    TkinterDnD = None
import requests
import fitz
import PyPDF2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# CAMINHOS DE CONFIGURAÇÃO
# Salva tudo em %APPDATA%\OrganizadorDocs  (Windows) ou ~/.OrganizadorDocs
# ══════════════════════════════════════════════════════════════════════════════

def _pasta_config() -> str:
    base = os.environ.get("APPDATA") or os.path.expanduser("~")
    pasta = os.path.join(base, "OrganizadorDocs")
    os.makedirs(pasta, exist_ok=True)
    return pasta

CONFIG_PATH  = os.path.join(_pasta_config(), "config.json")
CACHE_PATH   = os.path.join(_pasta_config(), "license_cache.json")

# Segredo HMAC local — impede que alguém edite o cache na mão
_HMAC_SECRET = b"OrganizadorDocs-v3-cache-secret"

SERVIDOR_SMTP = "smtp.gmail.com"
PORTA_SMTP    = 587
CACHE_DIAS    = 7   # dias que o programa funciona offline após última validação online


# ══════════════════════════════════════════════════════════════════════════════
# CONFIG.JSON  (e-mail SMTP, configurado pelo usuário na primeira execução)
# ══════════════════════════════════════════════════════════════════════════════

def carregar_config() -> dict:
    """Lê config.json; retorna dicionário vazio se não existir."""
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def salvar_config(dados: dict) -> None:
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=4)

def smtp_usuario() -> str:
    return carregar_config().get("smtp_usuario", "")

def smtp_senha() -> str:
    return carregar_config().get("smtp_senha", "")


# ══════════════════════════════════════════════════════════════════════════════
# TELA DE CONFIGURAÇÕES SMTP
# ══════════════════════════════════════════════════════════════════════════════

class JanelaConfiguracoes(tk.Toplevel):
    """Tela para o usuário cadastrar (ou alterar) o e-mail de envio."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Configurações de E-mail")
        self.resizable(False, False)
        self.grab_set()          # modal
        self._build()
        self._carregar()

    def _build(self):
        pad = dict(padx=12, pady=6)
        tk.Label(self, text="Conta de envio (Gmail)", font=("", 11, "bold")).grid(
            row=0, column=0, columnspan=2, pady=(14, 4), padx=12)

        tk.Label(self, text="E-mail:").grid(row=1, column=0, sticky="e", **pad)
        self._email = tk.Entry(self, width=36)
        self._email.grid(row=1, column=1, sticky="w", **pad)

        tk.Label(self, text="Senha de App:").grid(row=2, column=0, sticky="e", **pad)
        self._senha = tk.Entry(self, width=36, show="•")
        self._senha.grid(row=2, column=1, sticky="w", **pad)

        tk.Label(
            self,
            text="Use uma Senha de App do Google (não sua senha normal).\n"
                 "Gere em: Conta Google → Segurança → Senhas de app.",
            font=("", 8), fg="#666", justify="left",
            wraplength=320,
        ).grid(row=3, column=0, columnspan=2, padx=12, pady=(0, 8))

        # botões
        frm = tk.Frame(self)
        frm.grid(row=4, column=0, columnspan=2, pady=(4, 14))
        tk.Button(frm, text="Salvar", bg="#1a3d6e", fg="white",
                  relief="flat", width=10, command=self._salvar).pack(side="left", padx=6)
        tk.Button(frm, text="Cancelar", relief="flat", width=10,
                  command=self.destroy).pack(side="left", padx=6)

    def _carregar(self):
        cfg = carregar_config()
        self._email.insert(0, cfg.get("smtp_usuario", ""))
        self._senha.insert(0, cfg.get("smtp_senha", ""))

    def _salvar(self):
        email = self._email.get().strip()
        senha = self._senha.get().strip()
        if not email or not senha:
            messagebox.showwarning("Atenção", "Preencha e-mail e senha.", parent=self)
            return
        cfg = carregar_config()
        cfg["smtp_usuario"] = email
        cfg["smtp_senha"]   = senha
        salvar_config(cfg)
        messagebox.showinfo("Salvo", "Configurações salvas com sucesso!", parent=self)
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# SISTEMA DE LICENÇA  —  Google Sheets público + cache local HMAC (7 dias)
#
# Como configurar a planilha:
#   1. Crie uma Google Planilha com as chaves de licença (uma por linha, coluna A)
#   2. Publique: Arquivo → Compartilhar → Publicar na web → CSV → Planilha 1
#   3. Cole a URL gerada na constante LICENCA_SHEET_URL abaixo
# ══════════════════════════════════════════════════════════════════════════════

LICENCA_SHEET_URL = "https://docs.google.com/spreadsheets/d/1Ay_6ZZnV7LyufLCO6ZvYQhvBk-LHWgkEjiuhWDSdhoI/export?format=csv&gid=0"


def _assinar(payload: str) -> str:
    return hmac.new(_HMAC_SECRET, payload.encode(), hashlib.sha256).hexdigest()


def _cache_valido(chave: str) -> bool:
    """Retorna True se o cache local indica que a chave foi validada
    online nos últimos CACHE_DIAS dias e a assinatura HMAC bate."""
    try:
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            dados = json.load(f)
        if dados.get("chave") != chave:
            return False
        payload = f"{dados['chave']}|{dados['validade']}"
        if not hmac.compare_digest(_assinar(payload), dados.get("assinatura", "")):
            return False
        validade = datetime.fromisoformat(dados["validade"])
        return datetime.now(timezone.utc) < validade
    except Exception:
        return False


def _salvar_cache(chave: str) -> None:
    from datetime import timedelta
    validade = (datetime.now(timezone.utc) + timedelta(days=CACHE_DIAS)).isoformat()
    payload  = f"{chave}|{validade}"
    dados = {"chave": chave, "validade": validade, "assinatura": _assinar(payload)}
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(dados, f)


def _limpar_cache() -> None:
    try:
        os.remove(CACHE_PATH)
    except Exception:
        pass


def validar_licenca_online(chave: str) -> bool:
    """Consulta a planilha Google Sheets (CSV público)."""
    try:
        r = requests.get(LICENCA_SHEET_URL, timeout=6)
        r.raise_for_status()
        chaves = [linha.split(",")[0].strip() for linha in r.text.splitlines()]
        return chave in chaves
    except Exception:
        return False


def validar_licenca(chave: str) -> tuple[bool, str]:
    """
    Retorna (ok, mensagem).
    Fluxo:
      1. Cache local válido  → OK offline
      2. Online OK           → renova cache → OK
      3. Cache vencido + offline → BLOQUEADO
    """
    if _cache_valido(chave):
        return True, "offline_cache"

    ok_online = validar_licenca_online(chave)
    if ok_online:
        _salvar_cache(chave)
        return True, "online"

    # Sem internet e sem cache → bloqueia
    _limpar_cache()
    return False, "invalida"


# ══════════════════════════════════════════════════════════════════════════════
# TELA DE LICENÇA  (substitui o código morto que estava solto no módulo)
# ══════════════════════════════════════════════════════════════════════════════

class JanelaLicenca(tk.Tk):
    """Janela de validação de licença. Destrói a si mesma e sinaliza
    para o entry point se deve abrir o App."""

    def __init__(self):
        super().__init__()
        self.title("Organizador de Documentos — Licença")
        self.resizable(False, False)
        self.resultado = False   # True = licença OK
        self._build()
        # tenta usar chave já salva em config
        chave_salva = carregar_config().get("licenca", "")
        if chave_salva:
            self._entrada.insert(0, chave_salva)

    def _build(self):
        tk.Label(self, text="Organizador de Documentos Trabalhistas",
                 font=("", 11, "bold"), fg="#1a2e4a").pack(pady=(18, 4), padx=24)
        tk.Label(self, text="Digite sua chave de licença:",
                 font=("", 9)).pack()
        self._entrada = tk.Entry(self, width=38, font=("", 10))
        self._entrada.pack(pady=8, padx=24)
        self._status = tk.Label(self, text="", font=("", 8), fg="#c0392b")
        self._status.pack()
        tk.Button(self, text="Validar", bg="#1a3d6e", fg="white",
                  relief="flat", font=("", 10, "bold"), width=14,
                  command=self._verificar).pack(pady=(4, 18))
        self.bind("<Return>", lambda _: self._verificar())

    def _verificar(self):
        chave = self._entrada.get().strip()
        if not chave:
            self._status.config(text="Informe a chave de licença.")
            return
        self._status.config(text="Verificando...", fg="#334466")
        self.update_idletasks()

        ok, motivo = validar_licenca(chave)
        if ok:
            cfg = carregar_config()
            cfg["licenca"] = chave
            salvar_config(cfg)
            origem = "(online)" if motivo == "online" else "(cache offline)"
            self._status.config(text=f"✅ Licença válida {origem}", fg="#2e7d32")
            self.update_idletasks()
            self.after(800, self._abrir_app)
        else:
            self._status.config(text="❌ Chave inválida ou sem conexão.", fg="#c0392b")

    def _abrir_app(self):
        self.resultado = True
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS DE LOG E SMTP
# ══════════════════════════════════════════════════════════════════════════════

def log_envio(msg):
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    linha = f"[{timestamp}] {msg}"
    print(linha)
    with open("log_execucao.txt", "a", encoding="utf-8") as f:
        f.write(linha + "\n")

def nome_ascii(nome):
    return unicodedata.normalize("NFKD", nome).encode("ascii", "ignore").decode("ascii")

def ajustar_pastas_envio(pasta_documentos, caminho_json, nome_pasta_procurada):
    with open(caminho_json, 'r', encoding='utf-8') as f:
        clientes = json.load(f)

    pasta_temporario = os.path.join(pasta_documentos, "temporario")
    os.makedirs(pasta_temporario, exist_ok=True)

    padroes_pdf = [
        "boleto", "nota",
        "recibo1","recibo2","recibo3","recibo4",
        "recibo5","recibo6","recibo7","recibo8","recibo9"
    ]

    for cliente in clientes:
        condominio = cliente["condominio"]
        print(f"\n Processando condomínio: {condominio}")

        pasta_condominio = os.path.join(pasta_documentos, condominio)
        if not os.path.isdir(pasta_condominio):
            print(f" Pasta do condomínio não encontrada: {pasta_condominio}")
            continue

        pasta_mes_encontrada = None
        for ano in os.listdir(pasta_condominio):
            caminho_ano = os.path.join(pasta_condominio, ano)
            if not os.path.isdir(caminho_ano):
                continue
            caminho_mes = os.path.join(caminho_ano, nome_pasta_procurada)
            if os.path.isdir(caminho_mes):
                pasta_mes_encontrada = caminho_mes
                print(f" Pasta encontrada: {pasta_mes_encontrada}")
                break

        if not pasta_mes_encontrada:
            print(f" Pasta do mês '{nome_pasta_procurada}' não encontrada para {condominio}")
            continue

        pasta_destino = os.path.join(pasta_temporario, condominio)
        os.makedirs(pasta_destino, exist_ok=True)

        zip_path = os.path.join(pasta_destino, nome_pasta_procurada)
        shutil.make_archive(zip_path, 'zip', pasta_mes_encontrada)
        print(f" ZIP criado: {zip_path}.zip")

        for root, _, files in os.walk(pasta_mes_encontrada):
            for arquivo in files:
                if arquivo.lower().endswith(".pdf") and any(p in arquivo.lower() for p in padroes_pdf):
                    origem  = os.path.join(root, arquivo)
                    destino = os.path.join(pasta_destino, arquivo)
                    shutil.copy2(origem, destino)
                    print(f" PDF copiado: {arquivo}")

        cliente["caminho"] = pasta_destino

    with open(caminho_json, 'w', encoding='utf-8') as f:
        json.dump(clientes, f, ensure_ascii=False, indent=4)

def enviar_emails_faturamento(url_imagem, caminho_json, mes_competencia):
    relatorio   = []
    erros_envio = []

    with open(caminho_json, 'r', encoding='utf-8') as f:
        clientes = json.load(f)

    usuario = smtp_usuario()
    senha   = smtp_senha()
    if not usuario or not senha:
        raise RuntimeError(
            "E-mail de envio não configurado.\n"
            "Acesse o menu Configurações -> E-mail de Envio."
        )

    servidor = smtplib.SMTP(SERVIDOR_SMTP, PORTA_SMTP)
    servidor.starttls()
    try:
        servidor.login(usuario, senha)
    except Exception:
        servidor.quit()
        raise

    for cliente in tqdm(clientes, desc="Enviando e-mails", unit="email"):
        condominio    = cliente["condominio"]
        emails        = cliente["endereco"]
        caminho_pasta = cliente["caminho"]
        status = "✅ ENVIADO"
        motivo = ""

        try:
            arquivos = os.listdir(caminho_pasta)
            if not arquivos:
                raise Exception("Pasta vazia")

            destinatarios = [e.strip() for e in emails.split(";") if e.strip()]

            msg = MIMEMultipart()
            msg["From"]    = usuario
            msg["To"]      = ", ".join(destinatarios)
            msg["Subject"] = f"Faturamento competência {mes_competencia} - {condominio}"

            corpo_html = f"""
            <html>
                <body>
                    <p>
                        Boa tarde,<br><br>
                        Segue em anexo a documentação referente ao faturamento comp. {mes_competencia}.<br><br>
                        Atenciosamente.
                    </p>
                    <img src="{url_imagem}">
                </body>
            </html>
            """
            msg.attach(MIMEText(corpo_html, "html", "utf-8"))

            for arquivo in arquivos:
                caminho_arquivo = os.path.join(caminho_pasta, arquivo)
                if os.path.isfile(caminho_arquivo):
                    with open(caminho_arquivo, "rb") as f:
                        parte = MIMEBase("application", "octet-stream")
                        parte.set_payload(f.read())
                        encoders.encode_base64(parte)
                        filename_ascii = nome_ascii(arquivo)
                        filename_utf8  = encode_rfc2231(arquivo, "utf-8")
                        parte.add_header(
                            "Content-Disposition",
                            f'attachment; filename="{filename_ascii}"; filename*=utf-8\'\'{filename_utf8}'
                        )
                        msg.attach(parte)

            servidor.sendmail(usuario, destinatarios, msg.as_string())

        except Exception as e:
            status = "NÃO ENVIADO"
            motivo = str(e)
            erros_envio.append(cliente)

        relatorio.append({
            "Condomínio": condominio,
            "E-mails":    emails,
            "Status":     status,
            "Motivo":     motivo,
            "Data/Hora":  datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        })

    try:
        servidor.quit()
    except Exception:
        pass
    pd.DataFrame(relatorio).to_excel("relatorio_envios.xlsx", index=False)

    if erros_envio:
        with open("clientes_com_erro.json", "w", encoding="utf-8") as f:
            json.dump(erros_envio, f, ensure_ascii=False, indent=4)


# ══════════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ══════════════════════════════════════════════════════════════════════════════

STOPWORDS = {"DE","DA","DO","DOS","DAS"}

def normalizar_nome(nome):
    nome = unicodedata.normalize("NFD", nome.upper())
    nome = "".join(c for c in nome if unicodedata.category(c) != "Mn")
    nome = re.sub(r"[^A-Z\s]", "", nome)
    return [p for p in nome.split() if p not in STOPWORDS]

def score_match(nome_json, nome_pdf):
    pj, pp = normalizar_nome(nome_json), normalizar_nome(nome_pdf)
    if not pj or not pp: return 0
    s = 0
    if pj[0]  == pp[0]:  s += 20
    if pj[-1] == pp[-1]: s += 50
    for p in pp:
        if p in pj: s += 10
    if difflib.SequenceMatcher(None," ".join(pj)," ".join(pp)).ratio() > 0.8: s += 10
    return s

def encontrar_melhor_match(texto, funcionarios, limite=70):
    cands = []
    for item in funcionarios:
        for nome in item["funcionarios"]:
            s = score_match(nome, texto)
            if s >= limite: cands.append((s, nome, item["condominio"]))
    cands.sort(reverse=True, key=lambda x: x[0])
    return cands[0] if len(cands) == 1 else None

def convert_xls_to_xlsx(xls_path):
    df = pd.read_excel(xls_path, header=None, engine="xlrd")
    new_path = os.path.splitext(xls_path)[0] + ".converted.xlsx"
    df.to_excel(new_path, index=False, header=False, engine="openpyxl")
    return new_path

def create_json_from_excel(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb.active
    condo_data = []
    # min_col=3 -> row[0]=Nome(C) ... row[4]=CPF(G) ... row[12]=Cliente(O)
    for row in sheet.iter_rows(min_row=4, min_col=3, max_col=15, values_only=True):
        nome, cpf, cond = row[0], row[4], row[12]
        if not nome or not cond: continue
        ex = next((c for c in condo_data if c["condominio"] == cond), None)
        if ex:
            ex["funcionarios"].append(nome)
            ex["cpfs"].append(cpf)
        else:
            condo_data.append({"condominio": cond, "funcionarios": [nome], "cpfs": [cpf]})
    return condo_data

def copiar_conteudo_pasta(origem, destino):
    for item in os.listdir(origem):
        src = os.path.join(origem, item)
        dst = os.path.join(destino, item)
        if os.path.isdir(src):
            shutil.copytree(src, dst, dirs_exist_ok=True)
        else:
            shutil.copy2(src, dst)

def limpar_caminho_dnd(path):
    path = path.strip()
    if path.startswith("{") and path.endswith("}"): path = path[1:-1]
    return path

def cnpj_no_pdf_fitz(caminho, cnpj):
    try:
        doc = fitz.open(caminho)
        for pg in doc:
            if cnpj in pg.get_text():
                doc.close(); return True
        doc.close()
    except: pass
    return False


# ══════════════════════════════════════════════════════════════════════════════
# WIDGET: DROP ZONE
# ══════════════════════════════════════════════════════════════════════════════

class DropZone(tk.Frame):
    CN = "#f0f4f8"; CH = "#d0e8ff"; COK = "#d4edda"; BOK = "#5aaa6a"; CB = "#b0c4de"

    def __init__(self, parent, label, icon, modo="arquivo", filetypes=None, **kw):
        super().__init__(parent, **kw)
        self.modo = modo; self.filetypes = filetypes or []
        self.var  = tk.StringVar()
        self.var.trace_add("write", self._on_var)
        self._build(label, icon)
        if DND_AVAILABLE: self._reg_dnd()

    def _build(self, label, icon):
        self.config(bg=self.CN, relief="groove", bd=2, cursor="hand2")
        self.columnconfigure(1, weight=1)
        tk.Label(self, text=icon, font=("",16), bg=self.CN, fg="#5577aa"
                ).grid(row=0, column=0, rowspan=2, padx=(8,4), pady=6)
        self._lbl  = tk.Label(self, text=label, font=("",9,"bold"),
                               bg=self.CN, fg="#334466", anchor="w")
        self._lbl.grid(row=0, column=1, sticky="w", pady=(6,0))
        self._dica = tk.Label(self, text="Arraste ou clique para selecionar",
                               font=("",8), bg=self.CN, fg="#8899aa", anchor="w")
        self._dica.grid(row=1, column=1, sticky="w")
        self._ent  = tk.Entry(self, textvariable=self.var, font=("",8),
                               relief="flat", bg=self.CN, fg="#223344",
                               readonlybackground=self.CN, state="readonly")
        self._ent.grid(row=2, column=0, columnspan=2, sticky="ew", padx=6, pady=(0,4))
        for w in (self, self._lbl, self._dica, self._ent):
            w.bind("<Button-1>", self._click)
            w.bind("<Enter>",    lambda e: self._set(self.CH, self.CB) if not self.var.get() else None)
            w.bind("<Leave>",    lambda e: self._set(self.CN, self.CB) if not self.var.get() else None)

    def _reg_dnd(self):
        self.drop_target_register(DND_FILES)
        self.dnd_bind("<<Drop>>",      self._drop)
        self.dnd_bind("<<DragEnter>>", lambda e: self._set(self.CH, self.CB))
        self.dnd_bind("<<DragLeave>>", lambda e: self._set(self.CN, self.CB))

    def _drop(self, event):
        path = limpar_caminho_dnd(event.data)
        self._set(self.CN, self.CB)
        if (self.modo=="pasta" and os.path.isdir(path)) or (self.modo=="arquivo" and os.path.isfile(path)):
            self.var.set(path)

    def _click(self, _=None):
        p = filedialog.askdirectory() if self.modo=="pasta" else filedialog.askopenfilename(filetypes=self.filetypes)
        if p: self.var.set(p)

    def _on_var(self, *_):
        if self.var.get():
            self._set(self.COK, self.BOK)
            self._dica.config(text="OK  " + os.path.basename(self.var.get()))
        else:
            self._set(self.CN, self.CB)
            self._dica.config(text="Arraste ou clique para selecionar")

    def _set(self, bg, bd):
        self.config(bg=bg, highlightbackground=bd, highlightthickness=2, highlightcolor=bd)
        self._lbl.config(bg=bg); self._dica.config(bg=bg)
        self._ent.config(bg=bg, readonlybackground=bg)

    def get(self):   return self.var.get().strip()
    def set(self,v): self.var.set(v)
    def clear(self): self.var.set("")


# ══════════════════════════════════════════════════════════════════════════════
# WIDGET: PAINEL DE PROGRESSO (múltiplos processos, sem distorção)
# ══════════════════════════════════════════════════════════════════════════════

class PainelProgresso(tk.Frame):
    """
    Exibe N barras de progresso alinhadas, uma por processo.
    As barras são criadas dinamicamente e ficam sempre visíveis,
    eliminando o problema de distorção ao alternar entre processos.
    """
    COR = "#ccd8e8"

    PROCESSOS = [
        ("holerite",  "Holerite / Comprovante / Cartão Ponto"),
        ("fgts",      "FGTS"),
        ("nf",        "NF / Boleto / Recibo"),
        ("extrato",   "Extrato Mensal"),
        ("certidoes", "FGTS Gerais / Certidões"),
        ("mescla",    "Mescla de Pastas"),
    ]

    def __init__(self, parent, **kw):
        super().__init__(parent, bg=self.COR, **kw)
        self._bars  = {}
        self._lbls  = {}
        self._build()

    def _build(self):
        tk.Label(self, text="Progresso", font=("",9,"bold"),
                 bg=self.COR, fg="#334466").grid(row=0, column=0, columnspan=2,
                                                  sticky="w", padx=8, pady=(6,2))
        for i, (key, nome) in enumerate(self.PROCESSOS, start=1):
            tk.Label(self, text=nome+":", font=("",8), bg=self.COR,
                     fg="#445566", anchor="e", width=32
                    ).grid(row=i, column=0, sticky="e", padx=(8,4), pady=2)

            bar = ttk.Progressbar(self, mode="determinate", length=340)
            bar.grid(row=i, column=1, sticky="w", padx=(0,8), pady=2)

            lbl = tk.Label(self, text="Aguardando", font=("",8),
                           bg=self.COR, fg="#778899", anchor="w", width=30)
            lbl.grid(row=i, column=2, sticky="w", padx=(4,8), pady=2)

            self._bars[key] = bar
            self._lbls[key] = lbl

    def set(self, key, texto, valor=None, maximo=None):
        bar = self._bars.get(key)
        lbl = self._lbls.get(key)
        if not bar: return
        if maximo is not None: bar["maximum"] = maximo
        if valor  is not None: bar["value"]   = valor
        if lbl: lbl.config(text=texto)
        self.update_idletasks()

    def reset(self, key):
        self.set(key, "Aguardando", 0, 100)


# ══════════════════════════════════════════════════════════════════════════════
# JANELA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

_Base = TkinterDnD.Tk if DND_AVAILABLE else tk.Tk

class App(_Base):
    BG  = "#eef2f7"
    SEC = "#dde6f0"

    def __init__(self):
        super().__init__()
        self.title("Organizador de Documentos Trabalhistas v3.1")
        self.configure(bg=self.BG)
        self.minsize(860, 600)
        self.resizable(True, True)
        self._criar_menu()
        self._build()

    # ──────────────────────────────────────────────────────────────────────────
    # MENU SUPERIOR
    # ──────────────────────────────────────────────────────────────────────────

    def _criar_menu(self):
        barra = tk.Menu(self)
        menu_cfg = tk.Menu(barra, tearoff=0)
        menu_cfg.add_command(
            label="E-mail de Envio (SMTP)…",
            command=lambda: JanelaConfiguracoes(self)
        )
        menu_cfg.add_separator()
        menu_cfg.add_command(label="Sair", command=self.destroy)
        barra.add_cascade(label="Configurações", menu=menu_cfg)
        self.config(menu=barra)

    # ──────────────────────────────────────────────────────────────────────────
    # LAYOUT PRINCIPAL COM NOTEBOOK (2 ABAS)
    # ──────────────────────────────────────────────────────────────────────────

    def _build(self):
        if not DND_AVAILABLE:
            tk.Label(self, text="  Drag & Drop desativado — instale: pip install tkinterdnd2",
                     bg="#fff3cd", fg="#856404", font=("",8), anchor="w"
                    ).pack(fill="x")

        # Aviso se SMTP não estiver configurado
        if not smtp_usuario():
            tk.Label(
                self,
                text="  ⚠  E-mail de envio não configurado. Acesse Configurações -> E-mail de Envio.",
                bg="#fff3cd", fg="#856404", font=("", 8), anchor="w",
                cursor="hand2",
            ).pack(fill="x")

        tk.Label(self, text="Organizador de Documentos Trabalhistas",
                 font=("",14,"bold"), bg=self.BG, fg="#1a2e4a"
                ).pack(anchor="w", padx=16, pady=(10,2))
        tk.Label(self,
                 text="Configure ano/mês/saída e use as abas abaixo. "
                      "Arraste arquivos diretamente nas zonas indicadas.",
                 font=("",9), bg=self.BG, fg="#556"
                ).pack(anchor="w", padx=16, pady=(0,6))

        # Configurações gerais (fora das abas — sempre visível)
        self._secao_config()

        # Notebook com 2 abas
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12, pady=6)

        aba_proc   = tk.Frame(nb, bg=self.BG)
        aba_rel    = tk.Frame(nb, bg=self.BG)
        aba_envio  = tk.Frame(nb, bg=self.BG)
        nb.add(aba_proc,  text="  Processamento  ")
        nb.add(aba_rel,   text="  Relatório / Auditoria  ")
        nb.add(aba_envio, text="  Envio de Faturamento  ")

        # Aba processamento: canvas rolável
        canvas = tk.Canvas(aba_proc, bg=self.BG, highlightthickness=0)
        sb     = ttk.Scrollbar(aba_proc, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._inner = tk.Frame(canvas, bg=self.BG)
        win = canvas.create_window((0,0), window=self._inner, anchor="nw")
        self._inner.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",
            lambda e: canvas.itemconfig(win, width=e.width))
        self.bind_all("<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # Seções de processamento
        self._secao_holerite()
        self._secao_fgts_nf()        # FGTS + NF juntos (JSON compartilhado)
        self._secao_extrato()
        self._secao_certidoes()
        self._secao_mescla()

        # Painel de progresso (fixo na parte inferior da aba proc)
        self.progresso = PainelProgresso(aba_proc, relief="sunken", bd=1)
        self.progresso.pack(fill="x", side="bottom", padx=0, pady=0)

        # Aba relatório
        self._build_aba_relatorio(aba_rel)

        # Aba envio de faturamento
        self._build_aba_envio(aba_envio)

    # ──────────────────────────────────────────────────────────────────────────
    # HELPERS DE LAYOUT
    # ──────────────────────────────────────────────────────────────────────────

    def _bloco(self, titulo, cor="#1a2e4a"):
        outer = tk.Frame(self._inner, bg=self.BG)
        outer.pack(fill="x", padx=10, pady=(6,2))
        tk.Label(outer, text=titulo, font=("",10,"bold"),
                 bg=self.BG, fg=cor).pack(anchor="w", pady=(0,2))
        inner = tk.Frame(outer, bg=self.SEC, relief="groove", bd=1)
        inner.pack(fill="x")
        return inner

    def _grid2(self, p):
        g = tk.Frame(p, bg=self.SEC)
        g.pack(fill="x", padx=8, pady=4)
        g.columnconfigure(0, weight=1); g.columnconfigure(1, weight=1)
        return g

    def _grid3(self, p):
        g = tk.Frame(p, bg=self.SEC)
        g.pack(fill="x", padx=8, pady=4)
        g.columnconfigure(0,weight=1); g.columnconfigure(1,weight=1); g.columnconfigure(2,weight=1)
        return g

    def _grid4(self, p):
        g = tk.Frame(p, bg=self.SEC)
        g.pack(fill="x", padx=8, pady=4)
        for i in range(4): g.columnconfigure(i, weight=1)
        return g

    def _btn(self, parent, texto, cor, cmd):
        tk.Button(parent, text=texto, bg=cor, fg="white",
                  font=("",10,"bold"), relief="flat", cursor="hand2",
                  pady=5, command=cmd
                 ).pack(fill="x", padx=8, pady=(4,8))

    def _lbl_sec(self, parent, texto, cor="#334"):
        tk.Label(parent, text=texto, bg=self.SEC,
                 font=("",9,"bold"), fg=cor
                ).pack(anchor="w", padx=8, pady=(4,0))

    # ──────────────────────────────────────────────────────────────────────────
    # CONFIGURAÇÕES GERAIS
    # ──────────────────────────────────────────────────────────────────────────

    def _secao_config(self):
        f = tk.Frame(self, bg="#d5e3f0", relief="groove", bd=1)
        f.pack(fill="x", padx=12, pady=(4,2))
        row = tk.Frame(f, bg="#d5e3f0")
        row.pack(fill="x", padx=8, pady=8)

        tk.Label(row, text="Ano:", bg="#d5e3f0", font=("",9,"bold")).pack(side="left", padx=(4,2))
        self.ano_var = tk.StringVar()
        tk.Entry(row, textvariable=self.ano_var, width=6, font=("",10)).pack(side="left", padx=(0,14))

        tk.Label(row, text="Mês:", bg="#d5e3f0", font=("",9,"bold")).pack(side="left", padx=(0,2))
        self.mes_var = tk.StringVar()
        ttk.Combobox(row, textvariable=self.mes_var, state="readonly", width=5,
                     values=["01","02","03","04","05","06","07","08","09","10","11","12"]
                    ).pack(side="left", padx=(0,14))

        tk.Label(row, text="Pasta de saída:", bg="#d5e3f0", font=("",9,"bold")).pack(side="left", padx=(0,2))
        self.saida_var = tk.StringVar()
        tk.Entry(row, textvariable=self.saida_var, width=36, font=("",9)).pack(side="left", padx=(0,4))
        tk.Button(row, text="Selecionar", relief="flat", bg="#a0b8d0", cursor="hand2",
                  font=("",8),
                  command=lambda: self.saida_var.set(
                      filedialog.askdirectory() or self.saida_var.get())
                 ).pack(side="left")

    # ──────────────────────────────────────────────────────────────────────────
    # SEÇÃO: HOLERITE / COMPROVANTE / CARTÃO PONTO
    # ──────────────────────────────────────────────────────────────────────────

    def _secao_holerite(self):
        f = self._bloco("  Holerite  /  Comprovante  /  Cartão Ponto", "#1a3d2e")
        self._lbl_sec(f, "Planilha Excel — exportação do sistema de cartão ponto (obrigatória):")
        self.dz_excel = DropZone(f, "Planilha Excel", "S", filetypes=[("Excel","*.xls *.xlsx")], bg=self.SEC)
        self.dz_excel.pack(fill="x", padx=8, pady=4)

        self._lbl_sec(f, "PDFs — deixe vazio o que não tiver neste mês:")
        g = self._grid3(f)
        self.dz_holerite    = DropZone(g, "Holerite",    "H", filetypes=[("PDF","*.pdf")], bg=self.SEC)
        self.dz_comprovante = DropZone(g, "Comprovante", "C", filetypes=[("PDF","*.pdf")], bg=self.SEC)
        self.dz_cartao      = DropZone(g, "Cartão Ponto","P", filetypes=[("PDF","*.pdf")], bg=self.SEC)
        self.dz_holerite.grid(row=0,column=0,sticky="nsew",padx=4,pady=2)
        self.dz_comprovante.grid(row=0,column=1,sticky="nsew",padx=4,pady=2)
        self.dz_cartao.grid(row=0,column=2,sticky="nsew",padx=4,pady=2)

        self._btn(f, "Processar Holerite / Comprovante / Cartão Ponto", "#2d6a4f",
                  lambda: threading.Thread(target=self._run_holerite, daemon=True).start())

    # ──────────────────────────────────────────────────────────────────────────
    # SEÇÃO: FGTS + NF/BOLETO/RECIBO (JSON de CNPJ compartilhado)
    # ──────────────────────────────────────────────────────────────────────────

    def _secao_fgts_nf(self):
        f = self._bloco("  FGTS  e  Notas Fiscais / Boletos / Recibos", "#12325e")

        # JSON compartilhado
        self._lbl_sec(f, "JSON de CNPJs — usado pelo FGTS e pelo NF/Boleto/Recibo:")
        self.dz_cnpj_json = DropZone(f, "JSON de CNPJs (CNPJs.json)", "J",
                                      filetypes=[("JSON","*.json")], bg=self.SEC)
        self.dz_cnpj_json.pack(fill="x", padx=8, pady=4)

        sep = tk.Frame(f, bg="#b8cde0", height=1)
        sep.pack(fill="x", padx=8, pady=4)

        # FGTS
        self._lbl_sec(f, "FGTS — Relatório:")
        row_fgts = tk.Frame(f, bg=self.SEC)
        row_fgts.pack(fill="x", padx=8, pady=(2,0))
        tk.Label(row_fgts, text="Nome da subpasta FGTS:", bg=self.SEC, font=("",9)).pack(side="left")
        self.fgts_subpasta = tk.StringVar(value="FGTS")
        tk.Entry(row_fgts, textvariable=self.fgts_subpasta, width=16, font=("",9)).pack(side="left", padx=6)

        self.dz_fgts_pdf = DropZone(f, "PDF Relatório FGTS", "F",
                                     filetypes=[("PDF","*.pdf")], bg=self.SEC)
        self.dz_fgts_pdf.pack(fill="x", padx=8, pady=4)

        self._btn(f, "Processar FGTS", "#1d4e89",
                  lambda: threading.Thread(target=self._run_fgts, daemon=True).start())

        sep2 = tk.Frame(f, bg="#b8cde0", height=1)
        sep2.pack(fill="x", padx=8, pady=4)

        # NF / Boleto / Recibo
        self._lbl_sec(f, "NF / Boleto / Recibo — Pasta com os PDFs:")
        self.dz_nf_pasta = DropZone(f, "Pasta com os PDFs", "D", modo="pasta", bg=self.SEC)
        self.dz_nf_pasta.pack(fill="x", padx=8, pady=4)

        self._btn(f, "Processar NF / Boleto / Recibo", "#5a3e8e",
                  lambda: threading.Thread(target=self._run_nf_boleto, daemon=True).start())

    # ──────────────────────────────────────────────────────────────────────────
    # SEÇÃO: EXTRATO MENSAL (do contador)
    # ──────────────────────────────────────────────────────────────────────────

    def _secao_extrato(self):
        f = self._bloco("  Extrato Mensal  (pasta do contador)", "#3d2a00")
        self._lbl_sec(f,
            "Pasta recebida do contador: cada subpasta = nome do cliente, "
            "com 'Extrato Mensal.pdf' dentro.")
        g = self._grid2(f)
        self.dz_extrato_origem = DropZone(g, "Pasta do Contador (origem)", "D",
                                           modo="pasta", bg=self.SEC)
        self.dz_extrato_saida  = DropZone(g, "Pasta de Saída (output)", "D",
                                           modo="pasta", bg=self.SEC)
        self.dz_extrato_origem.grid(row=0,column=0,sticky="nsew",padx=4,pady=(8,4))
        self.dz_extrato_saida.grid( row=0,column=1,sticky="nsew",padx=4,pady=(8,4))
        self._btn(f, "Processar Extrato Mensal", "#7a5200",
                  lambda: threading.Thread(target=self._run_extrato, daemon=True).start())

    # ──────────────────────────────────────────────────────────────────────────
    # SEÇÃO: FGTS GERAIS / CERTIDÕES (distribuir para todas as pastas)
    # ──────────────────────────────────────────────────────────────────────────

    def _secao_certidoes(self):
        f = self._bloco("  FGTS Gerais e Certidões  — distribuir para todos os clientes", "#4a2060")
        self._lbl_sec(f,
            "Selecione as pastas de origem (FGTS, Gerais, Certidões) e a pasta de clientes (output). "
            "O conteúdo será copiado para ANO/MES/[subpasta] de cada cliente.")

        g = self._grid4(f)
        self.dz_cer_fgts    = DropZone(g, "Pasta FGTS\n(Guia + Comprovante)", "F", modo="pasta", bg=self.SEC)
        self.dz_cer_gerais  = DropZone(g, "Pasta Gerais\n(DCTFWeb / GPS)", "G", modo="pasta", bg=self.SEC)
        self.dz_cer_certs   = DropZone(g, "Pasta Certidões\n(CNDs)", "C", modo="pasta", bg=self.SEC)
        self.dz_cer_clientes= DropZone(g, "Pasta de Clientes\n(output)", "D", modo="pasta", bg=self.SEC)
        self.dz_cer_fgts.grid(   row=0,column=0,sticky="nsew",padx=4,pady=(8,4))
        self.dz_cer_gerais.grid( row=0,column=1,sticky="nsew",padx=4,pady=(8,4))
        self.dz_cer_certs.grid(  row=0,column=2,sticky="nsew",padx=4,pady=(8,4))
        self.dz_cer_clientes.grid(row=0,column=3,sticky="nsew",padx=4,pady=(8,4))

        self._btn(f, "Distribuir FGTS Gerais e Certidões para todos os clientes", "#5a3e8e",
                  lambda: threading.Thread(target=self._run_certidoes, daemon=True).start())

    # ──────────────────────────────────────────────────────────────────────────
    # SEÇÃO: MESCLAR PASTAS
    # ──────────────────────────────────────────────────────────────────────────

    def _secao_mescla(self):
        f = self._bloco("  Mesclar Pastas  (juntador.json)", "#5c2e00")
        g = self._grid2(f)
        self.dz_mescla_json  = DropZone(g, "JSON de mescla (juntador.json)", "J",
                                         filetypes=[("JSON","*.json")], bg=self.SEC)
        self.dz_mescla_pasta = DropZone(g, "Pasta base (output)", "D",
                                         modo="pasta", bg=self.SEC)
        self.dz_mescla_json.grid( row=0,column=0,sticky="nsew",padx=4,pady=(8,4))
        self.dz_mescla_pasta.grid(row=0,column=1,sticky="nsew",padx=4,pady=(8,4))
        self._btn(f, "Executar Mescla", "#7a4419",
                  lambda: threading.Thread(target=self._run_mescla, daemon=True).start())

    # ──────────────────────────────────────────────────────────────────────────
    # ABA: RELATÓRIO / AUDITORIA
    # ──────────────────────────────────────────────────────────────────────────

    def _build_aba_relatorio(self, parent):
        f = tk.Frame(parent, bg=self.BG)
        f.pack(fill="both", expand=True, padx=16, pady=12)

        tk.Label(f, text="Gerar Relatório de Auditoria",
                 font=("",12,"bold"), bg=self.BG, fg="#1a2e4a").pack(anchor="w")
        tk.Label(f,
                 text="Aponta condomínios organizados, documentos faltando por condomínio "
                      "e documentos faltando por funcionário. Exporta para Excel.",
                 font=("",9), bg=self.BG, fg="#556", wraplength=700, justify="left"
                ).pack(anchor="w", pady=(2,12))

        # Inputs
        frm = tk.Frame(f, bg=self.SEC, relief="groove", bd=1)
        frm.pack(fill="x", pady=(0,10))

        row1 = tk.Frame(frm, bg=self.SEC)
        row1.pack(fill="x", padx=8, pady=8)
        tk.Label(row1, text="Pasta output (clientes):", bg=self.SEC, font=("",9,"bold")).pack(side="left")
        self.rel_pasta_var = tk.StringVar()
        tk.Entry(row1, textvariable=self.rel_pasta_var, width=46, font=("",9)).pack(side="left", padx=6)
        tk.Button(row1, text="Selecionar", relief="flat", bg="#a0b8d0", cursor="hand2",
                  command=lambda: self.rel_pasta_var.set(
                      filedialog.askdirectory() or self.rel_pasta_var.get())
                 ).pack(side="left")

        row2 = tk.Frame(frm, bg=self.SEC)
        row2.pack(fill="x", padx=8, pady=(0,8))
        tk.Label(row2, text="Ano:", bg=self.SEC, font=("",9,"bold")).pack(side="left")
        self.rel_ano_var = tk.StringVar()
        tk.Entry(row2, textvariable=self.rel_ano_var, width=6, font=("",9)).pack(side="left", padx=(4,14))
        tk.Label(row2, text="Mês:", bg=self.SEC, font=("",9,"bold")).pack(side="left")
        self.rel_mes_var = tk.StringVar()
        ttk.Combobox(row2, textvariable=self.rel_mes_var, state="readonly", width=5,
                     values=["01","02","03","04","05","06","07","08","09","10","11","12"]
                    ).pack(side="left", padx=(4,14))
        tk.Label(row2, text="Salvar relatório em:", bg=self.SEC, font=("",9,"bold")).pack(side="left")
        self.rel_saida_var = tk.StringVar(value="relatorio_auditoria.xlsx")
        tk.Entry(row2, textvariable=self.rel_saida_var, width=30, font=("",9)).pack(side="left", padx=4)
        tk.Button(row2, text="Salvar como...", relief="flat", bg="#a0b8d0", cursor="hand2",
                  command=self._escolher_saida_rel).pack(side="left")

        tk.Button(f, text="Gerar Relatório Excel", bg="#1a3d2e", fg="white",
                  font=("",11,"bold"), relief="flat", cursor="hand2", pady=8,
                  command=lambda: threading.Thread(
                      target=self._run_relatorio, daemon=True).start()
                 ).pack(fill="x", pady=4)

        # Preview na tela
        tk.Label(f, text="Preview:", font=("",9,"bold"), bg=self.BG).pack(anchor="w", pady=(8,2))
        cols = ("Condomínio","Subpasta","Status","Detalhe")
        self._tree = ttk.Treeview(f, columns=cols, show="headings", height=14)
        for c in cols:
            self._tree.heading(c, text=c)
            self._tree.column(c, width=160 if c!="Detalhe" else 280)
        vsb = ttk.Scrollbar(f, orient="vertical", command=self._tree.yview)
        hsb = ttk.Scrollbar(f, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self._rel_status = tk.Label(f, text="", font=("",9), bg=self.BG, fg="#334")
        self._rel_status.pack(anchor="w", pady=4)

    def _escolher_saida_rel(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile="relatorio_auditoria.xlsx")
        if p: self.rel_saida_var.set(p)

    # ══════════════════════════════════════════════════════════════════════════
    # HELPERS COMUNS
    # ══════════════════════════════════════════════════════════════════════════

    def _get_periodo(self):
        ano   = self.ano_var.get().strip()
        mes   = self.mes_var.get().strip().zfill(2)
        saida = self.saida_var.get().strip() or "output"
        if not ano or not mes:
            raise ValueError("Informe o ano e o mês nas Configurações Gerais.")
        return ano, mes, saida

    def _salvar_pagina(self, page, cond, nome, tipo, ano, mes, saida):
        pasta = os.path.join(saida, str(cond), ano, f"{mes}.{ano}", tipo)
        os.makedirs(pasta, exist_ok=True)
        w = PyPDF2.PdfWriter(); w.add_page(page)
        with open(os.path.join(pasta, f"{tipo.replace(' ','_')}_{nome}.pdf"), "wb") as fh:
            w.write(fh)

    # ══════════════════════════════════════════════════════════════════════════
    # LÓGICA: HOLERITE / COMPROVANTE / CARTÃO PONTO
    # ══════════════════════════════════════════════════════════════════════════

    def _run_holerite(self):
        try:
            import re
            import unicodedata
            from rapidfuzz import fuzz, process

            def normalizar_nome(nome):
                nome = nome.upper().strip()

                # remove acentos
                nome = ''.join(
                    c for c in unicodedata.normalize('NFD', nome)
                    if unicodedata.category(c) != 'Mn'
                )

                # remove caracteres especiais
                nome = re.sub(r'[^A-Z\s]', ' ', nome)

                # remove espaços duplicados
                nome = re.sub(r'\s+', ' ', nome).strip()

                return nome

            def extrair_nome_comprovante(texto):
                """
                Extrai o nome do favorecido do comprovante.

                Cobre dois formatos:
                  - TED/DOC antigo: 'Nome: ... Agência'
                  - PIX: 'nome do recebedor/favorecido/beneficiário: ...',
                    onde o campo seguinte pode ser chave, CPF/CNPJ,
                    instituição etc. (não necessariamente 'agência').
                """

                texto = texto.upper()

                # PIX: nome do recebedor / favorecido / beneficiário
                m = re.search(
                    r'NOME\s+DO\s+(?:RECEBEDOR|FAVORECIDO|BENEFICI[ÁA]RIO):'
                    r'\s*([A-ZÀ-Ü\s]+?)\s*'
                    r'(?:\n|CHAVE|CPF|CNPJ|AG[EÊ]NCIA|CONTA|INSTITUI[ÇC][ÃA]O)',
                    texto
                )

                # TED/DOC antigo: 'Nome:' seguido de 'Agência'
                if not m:
                    m = re.search(
                        r'NOME:\s*([A-Z\s]+?)\s+AG[EÊ]NCIA',
                        texto,
                        re.DOTALL
                    )

                if not m:
                    return None

                nome = m.group(1)

                return normalizar_nome(nome)

            def extrair_cpf_comprovante(texto):
                """
                Extrai os dígitos visíveis do CPF do recebedor/favorecido
                em comprovantes PIX, que costumam vir parcialmente
                mascarados (ex.: '*****386881-**'). Usado como critério
                extra de match quando o nome vem abreviado/incompleto ou
                não bate direto com o cadastro.
                """

                texto_u = texto.upper()

                m = re.search(
                    r'CPF\s*/?\s*CNPJ\s+DO\s+(?:RECEBEDOR|FAVORECIDO|BENEFICI[ÁA]RIO):?'
                    r'\s*([\d\*\.\-\s]+?)'
                    r'(?:\n|INSTITUI[ÇC][ÃA]O|CHAVE|AG[EÊ]NCIA|CONTA|$)',
                    texto_u
                )

                if not m:
                    return None

                digitos = re.sub(r'\D', '', m.group(1))

                return digitos or None

            def normalizar_numero(bruto):
                """
                Normaliza um número de agência/conta: mantém só dígitos
                e remove zeros à esquerda, para que '0730' == '730' e
                '04964-2' == '4964-2'.
                """
                digitos = re.sub(r'\D', '', bruto or '')
                if not digitos:
                    return None
                return str(int(digitos))

            def extrair_agencia_conta(texto):
                """
                Extrai (agência, conta) tanto de holerites quanto de
                comprovantes bancários, cobrindo os formatos:
                  Holerite:    'Agência: 6729 - '   'conta: 18960-2'
                                                     'conta salário: 89288-9'
                  Comprovante: 'Agência: 7732'       'Conta corrente: 66890 - 5'
                """
                texto_u = texto.upper()

                # Em comprovantes há DUAS agências/contas: a da empresa
                # (conta debitada) e a do funcionário (conta creditada).
                # Restringe a busca ao trecho a partir de "CREDITADA"
                # quando esse marcador existir, para não pegar a conta
                # da empresa por engano.
                idx_creditada = texto_u.find('CREDITADA')
                trecho = texto_u[idx_creditada:] if idx_creditada != -1 else texto_u

                m_ag = re.search(r'AG[ÊE]NCIA:?\s*([0-9][0-9\.\s-]*)', trecho)
                m_ct = re.search(
                    r'CONTA(?:\s+SAL[ÁA]RIO|\s+CORRENTE)?:?\s*([0-9][0-9\.\s]*-\s*[0-9]+)',
                    trecho
                )

                if not m_ag or not m_ct:
                    return None

                agencia = normalizar_numero(m_ag.group(1))
                conta = normalizar_numero(m_ct.group(1))

                if not agencia or not conta:
                    return None

                return (agencia, conta)

            # 🔥 mapa (agência, conta) → (condomínio, nome_original), construído
            # a partir dos holerites à medida que forem processados/casados por
            # nome. Serve de segundo critério de match para os comprovantes,
            # quando o nome do favorecido no comprovante vem abreviado/incompleto.
            contas_map = {}

            ano, mes, saida = self._get_periodo()

            excel = self.dz_excel.get()

            if not excel:
                raise ValueError("Selecione a planilha Excel.")

            self.progresso.set(
                "holerite",
                "Carregando planilha...",
                0,
                100
            )

            # Converte XLS se necessário
            if excel.lower().endswith(".xls"):
                self.progresso.set(
                    "holerite",
                    "Convertendo XLS→XLSX..."
                )

                excel = convert_xls_to_xlsx(excel)

            funcionarios = create_json_from_excel(excel)

            # 🔥 mapa normalizado (agora carrega também o CPF, só com
            # dígitos, usado como critério extra de match nos comprovantes
            # PIX, onde o número vem parcialmente exposto)
            candidatos = {}

            for item in funcionarios:
                cond = item["condominio"]
                cpfs = item.get("cpfs", [])

                for idx, nome in enumerate(item["funcionarios"]):

                    if nome:
                        nome_norm = normalizar_nome(nome)

                        cpf_bruto = cpfs[idx] if idx < len(cpfs) else None
                        cpf_digitos = (
                            re.sub(r'\D', '', str(cpf_bruto))
                            if cpf_bruto else None
                        )

                        candidatos[nome_norm] = (
                            cond,
                            nome,
                            cpf_digitos
                        )

            arquivos = {
                t: dz.get()
                for t, dz in [
                    ("Holerites", self.dz_holerite),
                    ("Comprovantes", self.dz_comprovante),
                    ("Cartao Ponto", self.dz_cartao),
                ]
                if dz.get()
            }

            if not arquivos:
                raise ValueError("Selecione ao menos um PDF.")

            total = sum(
                len(PyPDF2.PdfReader(open(p, "rb")).pages)
                for p in arquivos.values()
            )

            self.progresso.set(
                "holerite",
                "Processando...",
                0,
                total
            )

            prog = 0

            for tipo, caminho in arquivos.items():

                with open(caminho, "rb") as fh:

                    reader = PyPDF2.PdfReader(fh)

                    for page in reader.pages:

                        texto = (page.extract_text() or "")
                        texto_upper = texto.upper()

                        encontrado = False

                        # ==================================================
                        # 🔹 COMPROVANTES
                        # ==================================================

                        if tipo == "Comprovantes":

                            nome_pdf = extrair_nome_comprovante(texto)

                            # 🔹 MATCH EXATO PRIMEIRO
                            if nome_pdf and nome_pdf in candidatos:

                                cond, nome_original, _cpf = candidatos[nome_pdf]

                                self._salvar_pagina(
                                    page,
                                    cond,
                                    nome_original,
                                    tipo,
                                    ano,
                                    mes,
                                    saida
                                )

                                encontrado = True

                            # 🔹 MATCH POR AGÊNCIA + CONTA (nome do banco
                            # abreviado/incompleto demais para o exato,
                            # mas a conta bate com a do holerite). Não
                            # existe em comprovantes PIX (não há agência/
                            # conta do recebedor), só em TED/DOC.
                            if not encontrado and contas_map:

                                chave_conta = extrair_agencia_conta(texto)

                                if chave_conta and chave_conta in contas_map:

                                    cond, nome_original = contas_map[chave_conta]

                                    self._salvar_pagina(
                                        page,
                                        cond,
                                        nome_original,
                                        tipo,
                                        ano,
                                        mes,
                                        saida
                                    )

                                    encontrado = True

                            # 🔹 MATCH POR CPF EXPOSTO (comprovantes PIX)
                            # O comprovante PIX mostra o CPF do recebedor
                            # parcialmente mascarado (ex.: '*****386881-**').
                            # Cruza o trecho exposto com o CPF completo de
                            # cada funcionário na planilha; como é muito
                            # improvável dois funcionários compartilharem o
                            # mesmo trecho exposto, exige-se ainda que o
                            # nome extraído (quando houver) seja parecido,
                            # como camada extra de segurança.
                            if not encontrado:

                                cpf_exposto = extrair_cpf_comprovante(texto)

                                if cpf_exposto:

                                    candidatos_cpf = [
                                        (nome_norm, cond, nome_original)
                                        for nome_norm, (cond, nome_original, cpf_digitos)
                                        in candidatos.items()
                                        if cpf_digitos and cpf_exposto in cpf_digitos
                                    ]

                                    if len(candidatos_cpf) == 1:

                                        nome_norm_cpf, cond, nome_original = candidatos_cpf[0]

                                        nome_ok = (
                                            not nome_pdf
                                            or fuzz.token_sort_ratio(
                                                nome_pdf, nome_norm_cpf
                                            ) >= 50
                                        )

                                        if nome_ok:

                                            self._salvar_pagina(
                                                page,
                                                cond,
                                                nome_original,
                                                tipo,
                                                ano,
                                                mes,
                                                saida
                                            )

                                            encontrado = True

                                    elif len(candidatos_cpf) > 1 and nome_pdf:

                                        # mais de um funcionário com o mesmo
                                        # trecho de CPF exposto (raríssimo):
                                        # desempata pelo nome
                                        nomes_cpf = [c[0] for c in candidatos_cpf]

                                        match = process.extractOne(
                                            nome_pdf,
                                            nomes_cpf,
                                            scorer=fuzz.token_sort_ratio,
                                            score_cutoff=70
                                        )

                                        if match:

                                            escolhido = next(
                                                c for c in candidatos_cpf
                                                if c[0] == match[0]
                                            )
                                            _, cond, nome_original = escolhido

                                            self._salvar_pagina(
                                                page,
                                                cond,
                                                nome_original,
                                                tipo,
                                                ano,
                                                mes,
                                                saida
                                            )

                                            encontrado = True

                            # 🔹 FUZZY MATCH (último recurso, precisa de
                            # nome extraído do comprovante)
                            if not encontrado and nome_pdf:

                                match = process.extractOne(
                                    nome_pdf,
                                    candidatos.keys(),
                                    scorer=fuzz.token_sort_ratio,
                                    score_cutoff=75
                                )

                                if match:

                                    nome_match = match[0]

                                    cond, nome_original, _cpf = candidatos[nome_match]

                                    self._salvar_pagina(
                                        page,
                                        cond,
                                        nome_original,
                                        tipo,
                                        ano,
                                        mes,
                                        saida
                                    )

                                    encontrado = True

                        # ==================================================
                        # 🔹 HOLERITES / CARTÃO PONTO
                        # ==================================================

                        else:

                            for nome_norm, (
                                cond,
                                nome_original,
                                _cpf
                            ) in candidatos.items():

                                if nome_norm in normalizar_nome(texto_upper):

                                    self._salvar_pagina(
                                        page,
                                        cond,
                                        nome_original,
                                        tipo,
                                        ano,
                                        mes,
                                        saida
                                    )

                                    encontrado = True

                                    # registra a conta bancária deste
                                    # funcionário para servir de match
                                    # alternativo nos comprovantes
                                    if tipo == "Holerites":
                                        chave_conta = extrair_agencia_conta(texto)
                                        if chave_conta:
                                            contas_map[chave_conta] = (
                                                cond,
                                                nome_original
                                            )

                                    break

                        # ==================================================
                        # 🔹 SEM MATCH
                        # ==================================================

                        if not encontrado:

                            self._salvar_pagina(
                                page,
                                "__PENDENTE",
                                "SEM_MATCH",
                                tipo,
                                ano,
                                mes,
                                saida
                            )

                        prog += 1

                        self.progresso.set(
                            "holerite",
                            f"{tipo} ({prog}/{total})",
                            prog,
                            total
                        )

            self.progresso.set(
                "holerite",
                "Concluído!",
                total,
                total
            )

            messagebox.showinfo(
                "Sucesso",
                "Holerite / Comprovante / Cartão Ponto\nconcluídos com sucesso!"
            )

        except Exception as e:

            self.progresso.set(
                "holerite",
                f"Erro: {e}"
            )

            messagebox.showerror(
                "Erro",
                str(e)
            )

    # ══════════════════════════════════════════════════════════════════════════
    # LÓGICA: FGTS
    # ══════════════════════════════════════════════════════════════════════════

    def _run_fgts(self):
        try:
            ano, mes, saida = self._get_periodo()
            subpasta  = self.fgts_subpasta.get().strip() or "FGTS"
            json_path = self.dz_cnpj_json.get()
            pdf_path  = self.dz_fgts_pdf.get()
            if not json_path or not pdf_path:
                raise ValueError("Selecione o JSON de CNPJs e o PDF do FGTS.")

            with open(json_path,"r",encoding="utf-8") as fh: data = json.load(fh)

            with open(pdf_path,"rb") as fh:
                reader = PyPDF2.PdfReader(fh)
                total_pgs = len(reader.pages)
                self.progresso.set("fgts","Processando...",0,len(data))

                for i,entry in enumerate(data):
                    cnpj, cond = entry["CNPJ"], entry["condominio"]
                    pasta = os.path.join(saida, cond, ano, f"{mes}.{ano}", subpasta)
                    os.makedirs(pasta, exist_ok=True)
                    pgs = [reader.pages[n] for n in range(total_pgs-1)
                           if cnpj in (reader.pages[n].extract_text() or "")]
                    if pgs:
                        w = PyPDF2.PdfWriter()
                        for pg in pgs: w.add_page(pg)
                        with open(os.path.join(pasta,"Relatorio FGTS Mensal.pdf"),"wb") as fw:
                            w.write(fw)
                    self.progresso.set("fgts", f"{i+1}/{len(data)} condomínios", i+1, len(data))

            self.progresso.set("fgts","Concluído!",len(data),len(data))
            messagebox.showinfo("Sucesso","Relatório FGTS organizado com sucesso!")
        except Exception as e:
            self.progresso.set("fgts",f"Erro: {e}")
            messagebox.showerror("Erro",str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # LÓGICA: NF / BOLETO / RECIBO
    # ══════════════════════════════════════════════════════════════════════════

    def _run_nf_boleto(self):
        try:
            ano, mes, saida = self._get_periodo()
            json_path  = self.dz_cnpj_json.get()
            pasta_docs = self.dz_nf_pasta.get()
            if not json_path or not pasta_docs:
                raise ValueError("Selecione o JSON de CNPJs e a pasta com PDFs.")

            with open(json_path,"r",encoding="utf-8") as fh:
                data = json.load(fh)

            # cria mapa CNPJ → condomínio
            mapa_cnpj = {d["CNPJ"]: d["condominio"] for d in data}

            pdfs = [a for a in os.listdir(pasta_docs) if a.lower().endswith(".pdf")]
            self.progresso.set("nf","Processando...",0,len(pdfs))

            for i, arq in enumerate(pdfs):
                cam = os.path.join(pasta_docs, arq)

                try:
                    doc = fitz.open(cam)
                    texto = ""
                    for pg in doc:
                        texto += pg.get_text()
                    doc.close()
                except:
                    continue

                # verifica qual CNPJ está no texto
                for cnpj, cond in mapa_cnpj.items():
                    if cnpj in texto:
                        nl = arq.lower()
                        sub = ("Boletos" if "boleto" in nl
                            else "Notas Fiscais" if "nota" in nl or "nf" in nl
                            else "Recibos" if "recibo" in nl
                            else "Documentos")

                        dst = os.path.join(saida, cond, ano, f"{mes}.{ano}", sub)
                        os.makedirs(dst, exist_ok=True)
                        shutil.move(cam, dst)
                        break

                self.progresso.set("nf", f"{i+1}/{len(pdfs)} PDFs", i+1, len(pdfs))

            self.progresso.set("nf","Concluído!",len(pdfs),len(pdfs))
            messagebox.showinfo("Sucesso","NF / Boleto / Recibo organizados com sucesso!")

        except Exception as e:
            self.progresso.set("nf",f"Erro: {e}")
            messagebox.showerror("Erro",str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # LÓGICA: EXTRATO MENSAL
    # ══════════════════════════════════════════════════════════════════════════

    def _run_extrato(self):
        try:
            ano = self.ano_var.get().strip()
            mes = self.mes_var.get().strip().zfill(2)
            if not ano or not mes:
                raise ValueError("Informe o ano e o mês nas Configurações Gerais.")

            origem = self.dz_extrato_origem.get()
            saida  = self.dz_extrato_saida.get() or self.saida_var.get().strip() or "output"
            if not origem: raise ValueError("Selecione a pasta do contador (origem).")

            clientes = [d for d in os.listdir(origem)
                        if os.path.isdir(os.path.join(origem, d))]
            self.progresso.set("extrato","Processando...",0,len(clientes))

            for i, cliente in enumerate(clientes):
                pasta_cliente_orig = os.path.join(origem, cliente)
                extratos = [f for f in os.listdir(pasta_cliente_orig)
                            if f.startswith("Extrato Mensal") and f.lower().endswith(".pdf")]
                if extratos:
                    dst = os.path.join(saida, cliente, ano, f"{mes}.{ano}", "Extrato Mensal")
                    os.makedirs(dst, exist_ok=True)
                    for arq in extratos:
                        src = os.path.join(pasta_cliente_orig, arq)
                        shutil.copy2(src, os.path.join(dst, arq))

                # Remove holerites soltos (como no script original)
                for arq in os.listdir(pasta_cliente_orig):
                    if arq.startswith("Recibo de Pagamento") and arq.lower().endswith(".pdf"):
                        try: os.remove(os.path.join(pasta_cliente_orig, arq))
                        except: pass

                self.progresso.set("extrato", f"{i+1}/{len(clientes)} clientes", i+1, len(clientes))

            self.progresso.set("extrato","Concluído!",len(clientes),len(clientes))
            messagebox.showinfo("Sucesso","Extrato Mensal organizado com sucesso!")
        except Exception as e:
            self.progresso.set("extrato",f"Erro: {e}")
            messagebox.showerror("Erro",str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # LÓGICA: FGTS GERAIS / CERTIDÕES → TODOS OS CLIENTES
    # ══════════════════════════════════════════════════════════════════════════

    def _run_certidoes(self):
        try:
            ano = self.ano_var.get().strip()
            mes = self.mes_var.get().strip().zfill(2)
            if not ano or not mes:
                raise ValueError("Informe o ano e o mês nas Configurações Gerais.")

            pasta_clientes = self.dz_cer_clientes.get()
            if not pasta_clientes:
                raise ValueError("Selecione a pasta de clientes (output).")

            # Mapeamento: subpasta destino → dropzone origem
            origens = {}
            if self.dz_cer_fgts.get():   origens["FGTS"]       = self.dz_cer_fgts.get()
            if self.dz_cer_gerais.get():  origens["Gerais"]     = self.dz_cer_gerais.get()
            if self.dz_cer_certs.get():   origens["Certidoes"]  = self.dz_cer_certs.get()

            if not origens:
                raise ValueError("Selecione ao menos uma pasta de documentos (FGTS, Gerais ou Certidões).")

            clientes = [d for d in os.listdir(pasta_clientes)
                        if os.path.isdir(os.path.join(pasta_clientes, d))]
            self.progresso.set("certidoes","Distribuindo...",0,len(clientes))

            for i, cliente in enumerate(clientes):
                for sub, origem in origens.items():
                    dst = os.path.join(pasta_clientes, cliente, ano, f"{mes}.{ano}", sub)
                    os.makedirs(dst, exist_ok=True)
                    for arq in os.listdir(origem):
                        src = os.path.join(origem, arq)
                        if os.path.isfile(src) and not arq.lower() == "desktop.ini":
                            shutil.copy2(src, os.path.join(dst, arq))
                self.progresso.set("certidoes", f"{i+1}/{len(clientes)} clientes", i+1, len(clientes))

            self.progresso.set("certidoes","Concluído!",len(clientes),len(clientes))
            messagebox.showinfo("Sucesso",
                                f"Documentos distribuídos para {len(clientes)} clientes com sucesso!")
        except Exception as e:
            self.progresso.set("certidoes",f"Erro: {e}")
            messagebox.showerror("Erro",str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # LÓGICA: MESCLAR PASTAS
    # ══════════════════════════════════════════════════════════════════════════

    def _run_mescla(self):
        try:
            json_path  = self.dz_mescla_json.get()
            pasta_base = self.dz_mescla_pasta.get()
            if not json_path or not pasta_base:
                raise ValueError("Selecione o JSON de mescla e a pasta base.")

            with open(json_path,"r",encoding="utf-8") as fh: dados = json.load(fh)
            self.progresso.set("mescla","Mesclando...",0,len(dados))
            deletadas = 0

            for i, item in enumerate(dados):
                destino = os.path.join(pasta_base, item["nome"])
                destino_norm = os.path.normcase(os.path.abspath(destino))
                os.makedirs(destino, exist_ok=True)
                for np_ in item["pastas"]:
                    origem = os.path.join(pasta_base, np_)
                    origem_norm = os.path.normcase(os.path.abspath(origem))
                    if origem_norm == destino_norm:
                        continue  # mesma pasta física (Windows ignora maiúsc/minúsc) - não copia nem apaga
                    if os.path.isdir(origem): copiar_conteudo_pasta(origem, destino)
                for np_ in item["pastas"]:
                    cam = os.path.join(pasta_base, np_)
                    cam_norm = os.path.normcase(os.path.abspath(cam))
                    if cam_norm == destino_norm:
                        continue  # não apagar a própria pasta de destino
                    if os.path.isdir(cam):
                        try: shutil.rmtree(cam); deletadas += 1
                        except Exception as ex: print(f"Erro ao deletar {cam}: {ex}")
                self.progresso.set("mescla", f"{i+1}/{len(dados)}", i+1, len(dados))

            self.progresso.set("mescla",f"Concluído! {deletadas} pastas removidas.",len(dados),len(dados))
            messagebox.showinfo("Sucesso",f"Mescla concluída!\n{deletadas} pastas removidas.")
        except Exception as e:
            self.progresso.set("mescla",f"Erro: {e}")
            messagebox.showerror("Erro",str(e))

    # ══════════════════════════════════════════════════════════════════════════
    # LÓGICA: RELATÓRIO DE AUDITORIA
    # ══════════════════════════════════════════════════════════════════════════

    # Subpastas esperadas por condomínio (nível do mês)
    SUBPASTAS_COND = ["Boletos","Certidoes","Extrato Mensal","FGTS",
                      "Gerais","Notas Fiscais","Recibos"]
    # Subpastas esperadas por funcionário
    SUBPASTAS_FUNC = ["Holerites","Comprovantes","Cartao Ponto"]

    def _run_relatorio(self):
        try:
            import os
            import unicodedata

            pasta_root = self.rel_pasta_var.get().strip()
            ano        = self.rel_ano_var.get().strip()
            mes        = self.rel_mes_var.get().strip().zfill(2)
            saida      = self.rel_saida_var.get().strip() or "relatorio_auditoria.xlsx"

            if not pasta_root:
                raise ValueError("Selecione a pasta de clientes.")

            if not ano or not mes:
                raise ValueError("Informe ano e mês.")

            periodo = f"{mes}.{ano}"

            self._rel_status.config(text="Analisando pastas...")
            self.update_idletasks()

            # Limpa preview
            for row in self._tree.get_children():
                self._tree.delete(row)

            linhas = []  # (condominio, subpasta, status, detalhe)

            condominios = sorted([
                d for d in os.listdir(pasta_root)
                if os.path.isdir(os.path.join(pasta_root, d))
            ])

            # ---------------------------------------------------
            # Função para padronizar nomes
            # ---------------------------------------------------
            def normalizar_nome(nome):
                nome = nome.lower().strip()

                nome = unicodedata.normalize('NFKD', nome)
                nome = ''.join(
                    c for c in nome
                    if not unicodedata.combining(c)
                )

                nome = nome.replace("_", " ")
                nome = " ".join(nome.split())

                return nome

            # ---------------------------------------------------
            # Prefixos possíveis dos PDFs
            # ---------------------------------------------------
            prefixos = [
                "Holerites_",
                "Comprovantes_",
                "Cartao_Ponto_",
                "CartaoPonto_",
                "Cartão_Ponto_",
                "CartãoPonto_",
            ]

            for cond in condominios:

                pasta_mes = os.path.join(
                    pasta_root,
                    cond,
                    ano,
                    periodo
                )

                if not os.path.isdir(pasta_mes):
                    linhas.append((
                        cond,
                        "—",
                        "SEM PASTA DO MÊS",
                        f"Não existe: {pasta_mes}"
                    ))
                    continue

                # ---------------------------------------------------
                # Verifica subpastas gerais
                # ---------------------------------------------------
                for sub in self.SUBPASTAS_COND:

                    pasta_sub = os.path.join(pasta_mes, sub)

                    if not os.path.isdir(pasta_sub):

                        linhas.append((
                            cond,
                            sub,
                            "FALTANDO",
                            "Subpasta não encontrada"
                        ))

                    else:

                        pdfs = [
                            f for f in os.listdir(pasta_sub)
                            if f.lower().endswith(".pdf")
                        ]

                        if not pdfs:

                            linhas.append((
                                cond,
                                sub,
                                "VAZIA",
                                "Nenhum PDF encontrado"
                            ))

                        else:

                            linhas.append((
                                cond,
                                sub,
                                "OK",
                                f"{len(pdfs)} arquivo(s)"
                            ))

                # ---------------------------------------------------
                # Verifica documentos dos funcionários
                # ---------------------------------------------------
                nomes_por_sub = {}

                for sub in self.SUBPASTAS_FUNC:

                    pasta_sub = os.path.join(pasta_mes, sub)

                    if not os.path.isdir(pasta_sub):
                        continue

                    arqs = [
                        os.path.splitext(f)[0]
                        for f in os.listdir(pasta_sub)
                        if f.lower().endswith(".pdf")
                    ]

                    nomes = set()

                    for a in arqs:

                        nome_limpo = a

                        # Remove prefixos conhecidos
                        for p in prefixos:
                            if nome_limpo.startswith(p):
                                nome_limpo = nome_limpo[len(p):]
                                break

                        nome_limpo = normalizar_nome(nome_limpo)

                        nomes.add(nome_limpo)

                    nomes_por_sub[sub] = nomes

                # ---------------------------------------------------
                # Compara funcionários entre pastas
                # ---------------------------------------------------
                if nomes_por_sub:

                    todos_nomes = set().union(*nomes_por_sub.values())

                    for nome in sorted(todos_nomes):

                        faltando = []

                        for s in self.SUBPASTAS_FUNC:

                            if s not in nomes_por_sub:
                                faltando.append(s)

                            elif nome not in nomes_por_sub[s]:
                                faltando.append(s)

                        if faltando:

                            linhas.append((
                                cond,
                                "Funcionário",
                                "INCOMPLETO",
                                f"{nome.upper()} — faltando: {', '.join(faltando)}"
                            ))

                        else:

                            linhas.append((
                                cond,
                                "Funcionário",
                                "OK",
                                nome.upper()
                            ))

            # ---------------------------------------------------
            # Atualiza preview
            # ---------------------------------------------------
            for row in linhas:

                tag = (
                    "ok" if row[2] == "OK"
                    else "falt" if row[2] in ("FALTANDO", "SEM PASTA DO MÊS")
                    else "inc" if row[2] in ("INCOMPLETO", "VAZIA")
                    else ""
                )

                self._tree.insert(
                    "",
                    "end",
                    values=row,
                    tags=(tag,)
                )

            self._tree.tag_configure(
                "ok",
                background="#e8f5e9"
            )

            self._tree.tag_configure(
                "falt",
                background="#fdecea"
            )

            self._tree.tag_configure(
                "inc",
                background="#fff8e1"
            )

            # ---------------------------------------------------
            # Exporta Excel
            # ---------------------------------------------------
            self._exportar_excel(
                linhas,
                saida,
                ano,
                mes
            )

            total_ok = sum(
                1 for l in linhas
                if l[2] == "OK"
            )

            total_prob = sum(
                1 for l in linhas
                if l[2] != "OK"
            )

            self._rel_status.config(
                text=(
                    f"Concluído — "
                    f"{total_ok} OK | "
                    f"{total_prob} problemas | "
                    f"Exportado: {saida}"
                )
            )

            messagebox.showinfo(
                "Relatório gerado",
                f"{len(condominios)} condomínios analisados.\n"
                f"{total_ok} itens OK | "
                f"{total_prob} problemas.\n"
                f"Arquivo: {saida}"
            )

        except Exception as e:

            self._rel_status.config(
                text=f"Erro: {e}"
            )

            messagebox.showerror(
                "Erro",
                str(e)
            )

    def _exportar_excel(self, linhas, caminho, ano, mes):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = f"Auditoria {mes}.{ano}"

        # Estilos
        h_fill = PatternFill("solid", fgColor="1A2E4A")
        h_font = Font(color="FFFFFF", bold=True, size=10)
        ok_fill   = PatternFill("solid", fgColor="C8E6C9")
        falt_fill = PatternFill("solid", fgColor="FFCDD2")
        inc_fill  = PatternFill("solid", fgColor="FFF9C4")
        borda = Border(
            left  =Side(style="thin"),right =Side(style="thin"),
            top   =Side(style="thin"),bottom=Side(style="thin"))

        headers = ["Condomínio","Subpasta / Categoria","Status","Detalhe"]
        ws.append(headers)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col)
            c.fill, c.font, c.alignment = h_fill, h_font, Alignment(horizontal="center")
            c.border = borda

        for row in linhas:
            ws.append(list(row))
            r = ws.max_row
            status = row[2]
            fill = (ok_fill if status=="OK"
                    else falt_fill if status in ("FALTANDO","SEM PASTA DO MÊS")
                    else inc_fill)
            for col in range(1, 5):
                c = ws.cell(row=r, column=col)
                c.fill, c.border = fill, borda
                c.alignment = Alignment(wrap_text=True)

        # Larguras
        for col, w in zip(range(1,5), [38, 22, 18, 50]):
            ws.column_dimensions[get_column_letter(col)].width = w

        # Aba resumo
        ws2 = wb.create_sheet("Resumo")
        condominios_unicos = sorted(set(l[0] for l in linhas))
        ws2.append(["Condomínio","Total Itens","OK","Problemas"])
        for cond in condominios_unicos:
            rows_c = [l for l in linhas if l[0]==cond]
            ok_c   = sum(1 for l in rows_c if l[2]=="OK")
            prob_c = sum(1 for l in rows_c if l[2]!="OK")
            ws2.append([cond, len(rows_c), ok_c, prob_c])
            r = ws2.max_row
            fill = ok_fill if prob_c == 0 else (falt_fill if prob_c > 2 else inc_fill)
            for col in range(1,5):
                ws2.cell(row=r,column=col).fill   = fill
                ws2.cell(row=r,column=col).border = borda

        for col, w in zip(range(1,5),[40,14,10,12]):
            ws2.column_dimensions[get_column_letter(col)].width = w

        # Cabeçalhos da aba resumo
        for col in range(1,5):
            c = ws2.cell(row=1,column=col)
            c.fill, c.font = h_fill, h_font
            c.border, c.alignment = borda, Alignment(horizontal="center")

        wb.save(caminho)



    # ══════════════════════════════════════════════════════════════════════════
    # ABA: ENVIO DE FATURAMENTO  (refatorada v3.2)
    # ══════════════════════════════════════════════════════════════════════════

    # Template padrão do corpo do e-mail (pode ser editado pelo usuário na UI)
    _TEMPLATE_PADRAO = (
        "Boa tarde,\n\n"
        "Segue em anexo a documentação referente ao faturamento comp. {competencia}.\n\n"
        "Atenciosamente."
    )

    def _build_aba_envio(self, parent):
        """
        Layout em dois painéis lado a lado (PanedWindow):
          • Esquerda  — configurações + editor de mensagem + botão de envio
          • Direita   — barra de tarefas (progresso por condomínio) + aba de log/e-mails
        """
        root = tk.Frame(parent, bg=self.BG)
        root.pack(fill="both", expand=True)

        paned = tk.PanedWindow(root, orient="horizontal", bg=self.BG,
                               sashwidth=6, sashrelief="flat", bd=0)
        paned.pack(fill="both", expand=True, padx=8, pady=8)

        # ── Painel esquerdo ──────────────────────────────────────────────────
        left = tk.Frame(paned, bg=self.BG)
        paned.add(left, minsize=420, stretch="always")

        tk.Label(left, text="Envio de Faturamento",
                 font=("", 12, "bold"), bg=self.BG, fg="#1a2e4a").pack(anchor="w", padx=4)
        tk.Label(left,
                 text="Ajusta pastas, gera ZIPs/PDFs e envia por e-mail.",
                 font=("", 9), bg=self.BG, fg="#556").pack(anchor="w", padx=4, pady=(0, 8))

        # — Bloco de configurações —
        cfg_frame = tk.LabelFrame(left, text=" Configurações ", bg=self.SEC,
                                   font=("", 9, "bold"), fg="#1a2e4a",
                                   relief="groove", bd=1)
        cfg_frame.pack(fill="x", padx=4, pady=(0, 6))

        def _cfg_row(parent, label, var, modo="entry", filetypes=None, width=38):
            row = tk.Frame(parent, bg=self.SEC)
            row.pack(fill="x", padx=8, pady=3)
            tk.Label(row, text=label, bg=self.SEC, font=("", 9, "bold"),
                     width=26, anchor="w").pack(side="left")
            ent = tk.Entry(row, textvariable=var, width=width, font=("", 9))
            ent.pack(side="left", padx=(0, 4))
            if modo == "pasta":
                tk.Button(row, text="📂", relief="flat", bg="#a0b8d0", cursor="hand2",
                          command=lambda v=var: v.set(filedialog.askdirectory() or v.get())
                         ).pack(side="left")
            elif modo == "arquivo":
                tk.Button(row, text="📄", relief="flat", bg="#a0b8d0", cursor="hand2",
                          command=lambda v=var, ft=filetypes: v.set(
                              filedialog.askopenfilename(filetypes=ft) or v.get())
                         ).pack(side="left")
            return ent

        self._env_pasta_var = tk.StringVar()
        self._env_json_var  = tk.StringVar()
        self._env_nomes_var = tk.StringVar()
        self._env_comp_var  = tk.StringVar()

        _cfg_row(cfg_frame, "Pasta de documentos:",    self._env_pasta_var, modo="pasta")
        _cfg_row(cfg_frame, "Arquivo JSON (clientes):", self._env_json_var,
                 modo="arquivo", filetypes=[("JSON", "*.json")])
        _cfg_row(cfg_frame, "Nome da pasta (mês):",    self._env_nomes_var, width=22)
        _cfg_row(cfg_frame, "Competência (ex: Jan 2026):", self._env_comp_var, width=22)

        # — Bloco de imagem (cartão de contato) —
        img_frame = tk.LabelFrame(left, text=" Cartão de Contato (imagem do e-mail) ",
                                   bg=self.SEC, font=("", 9, "bold"), fg="#1a2e4a",
                                   relief="groove", bd=1)
        img_frame.pack(fill="x", padx=4, pady=(0, 6))

        img_inner = tk.Frame(img_frame, bg=self.SEC)
        img_inner.pack(fill="x", padx=8, pady=6)

        self._env_img_modo = tk.StringVar(value="url")  # "url" ou "arquivo"

        rb_url = tk.Radiobutton(img_inner, text="URL", variable=self._env_img_modo,
                                value="url", bg=self.SEC, font=("", 9),
                                command=self._env_toggle_img)
        rb_url.pack(side="left", padx=(0, 4))
        rb_arq = tk.Radiobutton(img_inner, text="Arquivo local", variable=self._env_img_modo,
                                value="arquivo", bg=self.SEC, font=("", 9),
                                command=self._env_toggle_img)
        rb_arq.pack(side="left", padx=(0, 10))

        self._env_url_var  = tk.StringVar()
        self._env_img_path = tk.StringVar()

        self._env_img_url_ent = tk.Entry(img_inner, textvariable=self._env_url_var,
                                          width=38, font=("", 9))
        self._env_img_url_ent.pack(side="left", padx=(0, 4))

        self._env_img_file_frame = tk.Frame(img_inner, bg=self.SEC)
        self._env_img_file_ent = tk.Entry(self._env_img_file_frame,
                                           textvariable=self._env_img_path,
                                           width=30, font=("", 9))
        self._env_img_file_ent.pack(side="left", padx=(0, 4))
        tk.Button(self._env_img_file_frame, text="📎", relief="flat", bg="#a0b8d0",
                  cursor="hand2",
                  command=lambda: self._env_img_path.set(
                      filedialog.askopenfilename(
                          filetypes=[("Imagens", "*.png *.jpg *.jpeg *.gif *.bmp")])
                      or self._env_img_path.get())
                 ).pack(side="left")

        img_pos_row = tk.Frame(img_frame, bg=self.SEC)
        img_pos_row.pack(fill="x", padx=8, pady=(0, 6))
        tk.Label(img_pos_row, text="Posição da imagem:", bg=self.SEC, font=("", 9)).pack(side="left")
        self._env_img_pos = tk.StringVar(value="abaixo")
        for txt, val in [("Acima do texto", "acima"), ("Abaixo do texto", "abaixo")]:
            tk.Radiobutton(img_pos_row, text=txt, variable=self._env_img_pos,
                           value=val, bg=self.SEC, font=("", 9)).pack(side="left", padx=6)

        # — Editor de mensagem —
        msg_frame = tk.LabelFrame(left, text=" Mensagem do E-mail ",
                                   bg=self.SEC, font=("", 9, "bold"), fg="#1a2e4a",
                                   relief="groove", bd=1)
        msg_frame.pack(fill="both", expand=True, padx=4, pady=(0, 6))

        tk.Label(msg_frame,
                 text="Use {competencia} para inserir o mês automaticamente.",
                 font=("", 8), bg=self.SEC, fg="#668").pack(anchor="w", padx=8, pady=(4, 0))

        txt_frame = tk.Frame(msg_frame, bg=self.SEC)
        txt_frame.pack(fill="both", expand=True, padx=8, pady=4)
        self._env_msg_text = tk.Text(txt_frame, height=7, font=("", 9),
                                      relief="flat", bd=1, bg="white",
                                      wrap="word", undo=True)
        sb_msg = ttk.Scrollbar(txt_frame, orient="vertical",
                                command=self._env_msg_text.yview)
        self._env_msg_text.configure(yscrollcommand=sb_msg.set)
        self._env_msg_text.pack(side="left", fill="both", expand=True)
        sb_msg.pack(side="right", fill="y")
        self._env_msg_text.insert("1.0", self._TEMPLATE_PADRAO)

        btn_reset = tk.Button(msg_frame, text="↺ Restaurar padrão", relief="flat",
                               bg="#b8cde0", font=("", 8), cursor="hand2",
                               command=self._env_reset_msg)
        btn_reset.pack(anchor="e", padx=8, pady=(0, 6))

        # — Assunto —
        subj_row = tk.Frame(left, bg=self.BG)
        subj_row.pack(fill="x", padx=4, pady=(0, 6))
        tk.Label(subj_row, text="Assunto:", bg=self.BG,
                 font=("", 9, "bold"), width=10, anchor="w").pack(side="left")
        self._env_assunto_var = tk.StringVar(
            value="Faturamento competência {competencia} - {condominio}")
        tk.Entry(subj_row, textvariable=self._env_assunto_var,
                 width=52, font=("", 9)).pack(side="left")

        # — Botão principal —
        tk.Button(left, text="▶  Iniciar Processo", bg="#1a3d6e", fg="white",
                  font=("", 11, "bold"), relief="flat", cursor="hand2", pady=9,
                  command=lambda: threading.Thread(
                      target=self._run_envio, daemon=True).start()
                 ).pack(fill="x", padx=4, pady=(0, 4))

        self._env_status = tk.Label(left, text="", font=("", 9), bg=self.BG,
                                     fg="#334466", wraplength=400, justify="left")
        self._env_status.pack(anchor="w", padx=4)

        # ── Painel direito ───────────────────────────────────────────────────
        right = tk.Frame(paned, bg=self.BG)
        paned.add(right, minsize=340, stretch="always")

        # Sub-notebook: Progresso | Visualização de e-mails
        right_nb = ttk.Notebook(right)
        right_nb.pack(fill="both", expand=True, padx=4, pady=4)

        aba_prog  = tk.Frame(right_nb, bg=self.BG)
        aba_emails = tk.Frame(right_nb, bg=self.BG)
        right_nb.add(aba_prog,   text="  Progresso  ")
        right_nb.add(aba_emails, text="  E-mails Enviados  ")

        # — Aba de progresso: lista de condomínios —
        tk.Label(aba_prog, text="Status por condomínio",
                 font=("", 9, "bold"), bg=self.BG, fg="#1a2e4a").pack(
                     anchor="w", padx=8, pady=(8, 2))

        prog_cols = ("Condomínio", "Status")
        self._env_prog_tree = ttk.Treeview(aba_prog, columns=prog_cols,
                                            show="headings", height=20)
        self._env_prog_tree.heading("Condomínio", text="Condomínio")
        self._env_prog_tree.heading("Status",     text="Status")
        self._env_prog_tree.column("Condomínio",  width=220, anchor="w")
        self._env_prog_tree.column("Status",      width=100, anchor="center")

        # Tags de cor para os estados
        self._env_prog_tree.tag_configure("aguardando", foreground="#778899")
        self._env_prog_tree.tag_configure("enviando",   foreground="#1a3d6e", font=("", 9, "bold"))
        self._env_prog_tree.tag_configure("ok",         foreground="#2d6a4f", background="#eaf4ee")
        self._env_prog_tree.tag_configure("erro",       foreground="#c0392b", background="#fdecea")
        self._env_prog_tree.tag_configure("ajuste",     foreground="#7a5200", background="#fff8e8")

        vsb_prog = ttk.Scrollbar(aba_prog, orient="vertical",
                                  command=self._env_prog_tree.yview)
        self._env_prog_tree.configure(yscrollcommand=vsb_prog.set)
        self._env_prog_tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=4)
        vsb_prog.pack(side="right", fill="y", pady=4, padx=(0, 4))

        # barra global de progresso
        prog_bottom = tk.Frame(aba_prog, bg=self.BG)
        prog_bottom.pack(fill="x", padx=8, pady=(0, 6), side="bottom")
        self._env_prog_bar = ttk.Progressbar(prog_bottom, mode="determinate", length=300)
        self._env_prog_bar.pack(side="left", fill="x", expand=True)
        self._env_prog_lbl = tk.Label(prog_bottom, text="", font=("", 8), bg=self.BG, fg="#334")
        self._env_prog_lbl.pack(side="left", padx=6)

        # — Aba de visualização de e-mails enviados —
        tk.Label(aba_emails, text="E-mails desta sessão",
                 font=("", 9, "bold"), bg=self.BG, fg="#1a2e4a").pack(
                     anchor="w", padx=8, pady=(8, 2))

        email_cols = ("Condomínio", "Destinatários", "Status", "Hora")
        self._env_email_tree = ttk.Treeview(aba_emails, columns=email_cols,
                                             show="headings", height=8)
        self._env_email_tree.heading("Condomínio",   text="Condomínio")
        self._env_email_tree.heading("Destinatários", text="Destinatários")
        self._env_email_tree.heading("Status",        text="Status")
        self._env_email_tree.heading("Hora",          text="Hora")
        self._env_email_tree.column("Condomínio",     width=150, anchor="w")
        self._env_email_tree.column("Destinatários",  width=170, anchor="w")
        self._env_email_tree.column("Status",         width=80,  anchor="center")
        self._env_email_tree.column("Hora",           width=70,  anchor="center")
        self._env_email_tree.tag_configure("ok",   foreground="#2d6a4f", background="#eaf4ee")
        self._env_email_tree.tag_configure("erro", foreground="#c0392b", background="#fdecea")

        vsb_email = ttk.Scrollbar(aba_emails, orient="vertical",
                                   command=self._env_email_tree.yview)
        self._env_email_tree.configure(yscrollcommand=vsb_email.set)
        self._env_email_tree.pack(side="left", fill="both", expand=True, padx=(8, 0), pady=4)
        vsb_email.pack(side="right", fill="y", pady=4, padx=(0, 4))

        # Preview do e-mail selecionado
        tk.Label(aba_emails, text="Preview do e-mail selecionado:",
                 font=("", 8, "bold"), bg=self.BG, fg="#334").pack(
                     anchor="w", padx=8, pady=(4, 0))
        self._env_preview = tk.Text(aba_emails, height=8, font=("", 8),
                                     relief="flat", bd=1, bg="#f7f9fc",
                                     state="disabled", wrap="word")
        self._env_preview.pack(fill="both", expand=True, padx=8, pady=(2, 8))
        self._env_email_tree.bind("<<TreeviewSelect>>", self._env_preview_email)

        # guardar dados dos e-mails para o preview
        self._env_email_data: dict[str, dict] = {}

        # estado inicial da imagem
        self._env_toggle_img()

    # ── helpers da aba de envio ──────────────────────────────────────────────

    def _env_toggle_img(self):
        """Alterna entre campo URL e campo arquivo conforme o radio selecionado."""
        modo = self._env_img_modo.get()
        if modo == "url":
            self._env_img_file_frame.pack_forget()
            self._env_img_url_ent.pack(side="left", padx=(0, 4))
        else:
            self._env_img_url_ent.pack_forget()
            self._env_img_file_frame.pack(side="left")

    def _env_reset_msg(self):
        self._env_msg_text.delete("1.0", "end")
        self._env_msg_text.insert("1.0", self._TEMPLATE_PADRAO)

    def _env_preview_email(self, _=None):
        sel = self._env_email_tree.selection()
        if not sel:
            return
        iid = sel[0]
        dados = self._env_email_data.get(iid, {})
        corpo = dados.get("corpo", "")
        self._env_preview.config(state="normal")
        self._env_preview.delete("1.0", "end")
        self._env_preview.insert("1.0", corpo)
        self._env_preview.config(state="disabled")

    def _env_prog_set(self, cond_iid: str, status: str, tag: str):
        """Atualiza a linha de um condomínio na árvore de progresso (thread-safe via after)."""
        def _do():
            try:
                self._env_prog_tree.set(cond_iid, "Status", status)
                self._env_prog_tree.item(cond_iid, tags=(tag,))
                self._env_prog_tree.see(cond_iid)
            except Exception:
                pass
        self.after(0, _do)

    def _env_prog_global(self, valor: int, maximo: int, texto: str):
        def _do():
            self._env_prog_bar.configure(maximum=maximo, value=valor)
            self._env_prog_lbl.config(text=texto)
        self.after(0, _do)

    def _env_add_email(self, cond: str, destinatarios: str, ok: bool, hora: str, corpo: str):
        """Adiciona linha na aba de e-mails enviados (thread-safe)."""
        def _do():
            tag  = "ok" if ok else "erro"
            st   = "OK" if ok else "Erro"
            iid  = self._env_email_tree.insert(
                "", "end", values=(cond, destinatarios, st, hora), tags=(tag,))
            self._env_email_data[iid] = {"corpo": corpo}
        self.after(0, _do)

    # ── lógica de envio ──────────────────────────────────────────────────────

    def _run_envio(self):
        """
        Mesmo fluxo de envio original (ajustar_pastas_envio → enviar_emails_faturamento),
        agora com feedback visual por condomínio na barra de tarefas lateral.
        """
        try:
            pasta_documentos = self._env_pasta_var.get().strip()
            caminho_json     = self._env_json_var.get().strip()
            nome_pasta       = self._env_nomes_var.get().strip()
            mes_comp         = self._env_comp_var.get().strip()

            if not pasta_documentos: raise ValueError("Selecione a pasta de documentos.")
            if not caminho_json:     raise ValueError("Selecione o arquivo JSON.")
            if not nome_pasta:       raise ValueError("Informe o nome da pasta/mês.")
            if not mes_comp:         raise ValueError("Informe a competência (ex: Janeiro 2026).")

            # Resolve imagem
            if self._env_img_modo.get() == "url":
                url_imagem = self._env_url_var.get().strip()
                img_local  = None
            else:
                url_imagem = ""
                img_local  = self._env_img_path.get().strip() or None

            template_msg   = self._env_msg_text.get("1.0", "end").rstrip("\n")
            template_assunto = self._env_assunto_var.get().strip()
            posicao_img    = self._env_img_pos.get()  # "acima" ou "abaixo"

            self.after(0, lambda: self._env_status.config(text="Carregando clientes..."))

            with open(caminho_json, "r", encoding="utf-8") as f:
                clientes = json.load(f)

            # Popula a árvore de progresso com todos os condomínios
            def _init_tree():
                for item in self._env_prog_tree.get_children():
                    self._env_prog_tree.delete(item)
                for cl in clientes:
                    self._env_prog_tree.insert(
                        "", "end",
                        iid=cl["condominio"],
                        values=(cl["condominio"], " Aguardando"),
                        tags=("aguardando",))
            self.after(0, _init_tree)

            log_envio("=" * 60)
            log_envio("INICIO DA EXECUCAO")
            log_envio(f"Pasta documentos : {pasta_documentos}")
            log_envio(f"Arquivo JSON     : {caminho_json}")
            log_envio(f"Pasta mês        : {nome_pasta}")
            log_envio(f"Competência      : {mes_comp}")

            # ── Fase 1: ajuste de pastas ─────────────────────────────────────
            self.after(0, lambda: self._env_status.config(
                text="[AJUSTE] Ajustando pastas e gerando ZIPs..."))
            for cl in clientes:
                self._env_prog_set(cl["condominio"], " Ajustando...", "ajuste")

            log_envio("[AJUSTE] Inicio do ajuste de pastas")
            ajustar_pastas_envio(pasta_documentos, caminho_json, nome_pasta)
            log_envio("[AJUSTE] Ajuste finalizado")

            # Recarrega JSON com caminhos atualizados
            with open(caminho_json, "r", encoding="utf-8") as f:
                clientes = json.load(f)

            # ── Fase 2: envio individual com feedback ────────────────────────
            self.after(0, lambda: self._env_status.config(text="[ENVIO] Enviando e-mails..."))
            log_envio("[ENVIO] Inicio do envio de emails")

            usuario = smtp_usuario()
            senha   = smtp_senha()
            if not usuario or not senha:
                raise RuntimeError(
                    "E-mail de envio não configurado.\n"
                    "Acesse Configurações -> E-mail de Envio.")

            servidor = smtplib.SMTP(SERVIDOR_SMTP, PORTA_SMTP)
            servidor.starttls()
            try:
                servidor.login(usuario, senha)
            except Exception:
                servidor.quit()
                raise

            relatorio   = []
            erros_envio = []
            total       = len(clientes)

            for idx, cliente in enumerate(clientes, 1):
                condominio    = cliente["condominio"]
                emails_str    = cliente["endereco"]
                caminho_pasta = cliente.get("caminho", "")
                status_rel    = " ENVIADO"
                motivo        = ""

                self._env_prog_set(condominio, "📤 Enviando...", "enviando")
                self._env_prog_global(idx - 1, total,
                                      f"{idx - 1}/{total}")

                assunto = (template_assunto
                           .replace("{competencia}", mes_comp)
                           .replace("{condominio}",  condominio))
                corpo_txt = template_msg.replace("{competencia}", mes_comp)

                # Monta HTML com posição de imagem configurável
                img_tag = ""
                if url_imagem:
                    img_tag = f'<img src="{url_imagem}" style="max-width:100%;">'
                elif img_local and os.path.isfile(img_local):
                    import base64, mimetypes
                    mt = mimetypes.guess_type(img_local)[0] or "image/png"
                    with open(img_local, "rb") as fimg:
                        b64 = base64.b64encode(fimg.read()).decode()
                    img_tag = (f'<img src="data:{mt};base64,{b64}" '
                               f'style="max-width:100%;">')

                corpo_html_lines = corpo_txt.replace("\n", "<br>")
                if posicao_img == "acima":
                    html_body = f"{img_tag}<br>{corpo_html_lines}"
                else:
                    html_body = f"{corpo_html_lines}<br><br>{img_tag}"

                corpo_html = (
                    "<html><body style='font-family:Arial,sans-serif;font-size:14px;'>"
                    f"<p>{html_body}</p>"
                    "</body></html>"
                )

                try:
                    arquivos = os.listdir(caminho_pasta) if caminho_pasta else []
                    if not arquivos:
                        raise Exception("Pasta vazia ou não encontrada")

                    destinatarios = [e.strip() for e in emails_str.split(";") if e.strip()]

                    msg = MIMEMultipart()
                    msg["From"]    = usuario
                    msg["To"]      = ", ".join(destinatarios)
                    msg["Subject"] = assunto
                    msg.attach(MIMEText(corpo_html, "html", "utf-8"))

                    for arquivo in arquivos:
                        caminho_arquivo = os.path.join(caminho_pasta, arquivo)
                        if os.path.isfile(caminho_arquivo):
                            with open(caminho_arquivo, "rb") as fa:
                                parte = MIMEBase("application", "octet-stream")
                                parte.set_payload(fa.read())
                                encoders.encode_base64(parte)
                                filename_ascii = nome_ascii(arquivo)
                                filename_utf8  = encode_rfc2231(arquivo, "utf-8")
                                parte.add_header(
                                    "Content-Disposition",
                                    f'attachment; filename="{filename_ascii}";'
                                    f' filename*=utf-8\'\'{filename_utf8}'
                                )
                                msg.attach(parte)

                    servidor.sendmail(usuario, destinatarios, msg.as_string())
                    self._env_prog_set(condominio, " Enviado", "ok")
                    log_envio(f"[OK] {condominio} -> {emails_str}")

                except Exception as e:
                    status_rel = "NÃO ENVIADO"
                    motivo     = str(e)
                    erros_envio.append(cliente)
                    self._env_prog_set(condominio, "Erro", "erro")
                    log_envio(f"[ERRO] {condominio}: {e}")

                hora_envio = datetime.now().strftime("%H:%M:%S")
                self._env_add_email(
                    condominio, emails_str,
                    ok=(status_rel == "ENVIADO"),
                    hora=hora_envio,
                    corpo=f"Assunto: {assunto}\n\n{corpo_txt}\n\n"
                          f"[Destinatários: {emails_str}]\n"
                          f"[Anexos: {len(arquivos)} arquivo(s)]"
                          + (f"\n\n⚠ Erro: {motivo}" if motivo else "")
                )

                relatorio.append({
                    "Condomínio": condominio,
                    "E-mails":    emails_str,
                    "Status":     status_rel,
                    "Motivo":     motivo,
                    "Data/Hora":  datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                })

            try:
                servidor.quit()
            except Exception:
                pass

            self._env_prog_global(total, total, f"{total}/{total}")
            pd.DataFrame(relatorio).to_excel("relatorio_envios.xlsx", index=False)

            if erros_envio:
                with open("clientes_com_erro.json", "w", encoding="utf-8") as f:
                    json.dump(erros_envio, f, ensure_ascii=False, indent=4)

            # ── Fase 3: limpeza ──────────────────────────────────────────────
            pasta_temporario = os.path.join(pasta_documentos, "temporario")
            if os.path.isdir(pasta_temporario):
                shutil.rmtree(pasta_temporario)
                log_envio("[LIMPEZA] Pasta 'temporario' removida")

            log_envio("[ENVIO] Envio finalizado")
            log_envio("FIM DA EXECUCAO - SUCESSO")

            enviados = sum(1 for r in relatorio if r["Status"] == "ENVIADO")
            erros    = total - enviados
            resumo   = (f" Processo finalizado!  "
                        f"{enviados}/{total} enviados"
                        + (f"  •  {erros} erro(s)" if erros else "")
                        + "  •  Veja 'relatorio_envios.xlsx'")
            self.after(0, lambda: self._env_status.config(text=resumo))
            messagebox.showinfo(
                "Concluído",
                f"Processo finalizado!\n\n"
                f"Enviados: {enviados}/{total}\n"
                f"Erros:    {erros}\n\n"
                "Relatório -> relatorio_envios.xlsx\n"
                "Log       -> log_execucao.txt"
            )

        except Exception as e:
            log_envio(f"[ERRO FATAL] {e}")
            self.after(0, lambda: self._env_status.config(text=f"❌ Erro: {e}"))
            messagebox.showerror("Erro", str(e))


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # 1. Valida licença (JanelaLicenca é um tk.Tk próprio)
    tela_licenca = JanelaLicenca()
    tela_licenca.mainloop()

    # 2. Só abre o App se a licença foi aprovada
    if tela_licenca.resultado:
        app = App()
        app.mainloop()